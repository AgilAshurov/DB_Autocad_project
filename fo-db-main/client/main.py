from startup import app_wd
from utils import Cell, Table, reload_combobox, tr, to_widget_ts, cable_type
import os
import sys
from PySide6.QtWidgets import QMainWindow, QApplication, QInputDialog, QLineEdit, QMessageBox, QListWidgetItem, QFileDialog
from PySide6.QtCore import QTranslator, QLibraryInfo, QLocale, Qt
from PySide6.QtGui import QAction, QPixmap
from enter_window_ui import Ui_EnterWindow
from main_window_ui import Ui_MainWindow
from item_types_window_ui import Ui_ItemTypesWindow
from users_window_ui import Ui_UsersWindow
from history_window_ui import Ui_HistoryWindow
from user_history_window_ui import Ui_UserHistoryWindow
from select_users_window_ui import Ui_SelectUsersWindow
import ctypes
import warnings
import json
from remote_backend import Backend
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
from loader import load
import traceback
import time
import re


backend = None
item_types = None
item_type_names = None
olts = None
users = None
history = None

enter_window = None
main_window = None
item_types_window = None
users_window = None
history_window = None
user_history_window = None
select_users_window = None

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

try:
    with open(os.path.join(app_wd, "settings.json"), "rb") as f:
        settings = json.loads(f.read().decode("UTF-8"))
except:
    settings = {}
settings.setdefault("language", "en")
settings.setdefault("url", "https://127.0.0.1:8080")
settings.setdefault("name", "")
settings.setdefault("timeout", 10)


def logout():
    global enter_window, main_window, item_types_window, users_window, history_window, user_history_window, select_users_window

    enter_window = EnterWindow()
    enter_window.show()

    if main_window is not None:
        main_window.close()
        main_window = None

    if item_types_window is not None:
        item_types_window.close()
        item_types_window = None

    if users_window is not None:
        users_window.close()
        users_window = None

    if history_window is not None:
        history_window.close()
        history_window = None

    if user_history_window is not None:
        user_history_window.close()
        user_history_window = None

    if select_users_window is not None:
        select_users_window.close()
        select_users_window = None


def danger(parent, title, text, buttons=QMessageBox.Ok):
    msg = QMessageBox(QMessageBox.NoIcon, title, text, buttons, parent)
    msg.setIconPixmap(QPixmap(":/icons/error.svg"))
    if parent is None:
        msg.setWindowIcon(QPixmap(":/app.png"))
    return msg.exec()


def repeat():
    if QMessageBox.critical(None, tr("Error"), tr("Request error. Retry request?"), QMessageBox.Retry | QMessageBox.Abort) == QMessageBox.Retry:
        return True
    logout()
    return False


def error():
    QMessageBox.critical(None, tr("Error"), tr("Server error."))
    logout()


class EnterWindow(QMainWindow, Ui_EnterWindow):
    LANGUAGES = [
        ("English", "en"),
        ("Русский", "ru"),
        ("Azərbaycan", "az")
    ]

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        language_index = 0
        for i, item in enumerate(self.LANGUAGES):
            name, language = item
            self.language.addItem(name, language)
            if language == settings["language"]:
                language_index = i
        self.language.setCurrentIndex(language_index)

        self.url.setText(settings["url"])
        self.name.setText(settings["name"])

        self.language.currentTextChanged.connect(self.language_changed)
        self.enter.clicked.connect(self.enter_clicked)
        self.cancel.clicked.connect(self.cancel_clicked)

        self.url.returnPressed.connect(self.enter_clicked)
        self.name.returnPressed.connect(self.enter_clicked)
        self.password.returnPressed.connect(self.enter_clicked)

    def enter_clicked(self):
        global backend, item_types, item_type_names, olts, main_window

        with open(os.path.join(app_wd, "settings.json"), "wb") as f:
            settings.update({
                "url": self.url.text(),
                "name": self.name.text()
            })
            f.write(json.dumps(settings).encode("UTF-8"))

        backend = Backend(settings["url"], settings["timeout"])

        sid = backend.enter(self.name.text(), self.password.text())
        if sid is None:
            QMessageBox.critical(self, tr("Error"), tr("Server error."))
            return
        if sid is False:
            danger(self, tr("Error"), tr("Bad user name or password."))
            return

        item_types = backend.select("item_type", None, "id")
        if item_types is False:
            QMessageBox.critical(self, tr("Error"), tr("Server error."))
            return

        item_type_names = [item_type["id"] for item_type in item_types]

        olts = backend.select("olt", None, "name")
        if olts is False:
            QMessageBox.critical(self, tr("Error"), tr("Server error."))
            return

        backend.repeat = repeat
        backend.error = error

        main_window = MainWindow()
        main_window.show()

        self.close()

    def cancel_clicked(self):
        self.close()

    def language_changed(self):
        set_language(self.language.currentData())
        self.retranslateUi(self)


class MainWindow(QMainWindow, Ui_MainWindow):
    def object_calculate(self, table, row, column=None):
        if column is not None and self.__objects.cells[column].name != "blocks":
            return
        obj = self.__objects.selected_row()
        if obj is not None and obj.get_row() == row:
            blocks = len(self.__items.cells) - len(self.ITEM_CELLS_PREFIX) - len(self.ITEM_CELLS_SUFFIX)
            if obj.blocks != blocks:
                self.reload_items(obj)
                self.reload_cables(obj)

    def item_calculate(self, table, row, column=None):
        if column is not None:
            name = self.__items.cells[column].name
            if not (name.startswith("block_") or name in ["basement", "project"]):
                return
        obj = self.__objects.selected_row()
        item = self.__items.row(row)
        item.total = sum([
            getattr(item, "block_{}".format(i + 1))
            for i in range(obj.blocks)
        ]) + item.basement
        item.reserved = item.project - item.total

    def cable_calculate(self, table, row, column=None):
        if column is not None:
            name = self.__cables.cells[column].name
            if not (name.startswith("block_") or name in ["basement", "project"]):
                return
        obj = self.__objects.selected_row()
        cable = self.__cables.row(row)
        cable.total = sum([
            getattr(cable, "block_{}".format(i + 1))
            for i in range(obj.blocks)
        ]) + cable.basement
        cable.reserved = cable.project - cable.total

    def reload_items(self, obj):
        self.ITEM_CELLS = (
            self.ITEM_CELLS_PREFIX +
            [Cell("i", "block_{}".format(i + 1), str(i + 1)) for i in range(obj.blocks)] +
            self.ITEM_CELLS_SUFFIX
        )
        self.__items = Table(self.items, self.ITEM_CELLS, [], self.item_calculate, backend, "item", self.on_create_editor)
        self.__items.name_to_cell["name"].items = item_type_names

        items = backend.select("item", {"object_id": obj.get_id()}, "name")
        if items is False:
            return
        self.__items.reload(items, self.show_deleted.isChecked(), not backend.user["editor"] and not backend.user["advanced_editor"])

    def reload_cables(self, obj):
        self.CABLE_CELLS = (
            self.CABLE_CELLS_PREFIX +
            [Cell("f", "block_{}".format(i + 1), str(i + 1)) for i in range(obj.blocks)] +
            self.CABLE_CELLS_SUFFIX
        )
        self.__cables = Table(self.cables, self.CABLE_CELLS, [], self.cable_calculate, backend, "cable", self.on_create_editor)

        cables = backend.select("cable", {"object_id": obj.get_id()}, "name")
        if cables is False:
            return
        self.__cables.reload(cables, self.show_deleted.isChecked(), not backend.user["editor"] and not backend.user["advanced_editor"])

    def on_create_editor(self, table, row, column, row_name):
        self.active_table = table
        self.active_row = row
        self.active_column = column
        self.active_row_name = row_name

    def validate_object(self, row, cell, value):
        if cell.column > 0:
            return True
        if value == "":
            QMessageBox.warning(self, tr("Warning"), tr("Error"))
            return False
        index = self.__objects.find(value)
        if index != -1 and index != row:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return False
        return True

    def validate_item(self, row, cell, value):
        if cell.column > 0:
            return True
        index = self.__items.find(value)
        if index != -1 and index != row:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return False
        return True

    def validate_cable(self, row, cell, value):
        if cell.column > 0:
            return True
        if cable_type(value) is None:
            QMessageBox.warning(self, tr("Warning"), tr("Use a cable type such as FO24 as a prefix."))
            return False
        index = self.__cables.find(value)
        if index != -1 and index != row:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return False
        return True

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.OLT_CELLS = [
            Cell("i", "gpon", tr("GPON-card x16")),
            Cell("i", "odf", tr("ODF x24")),
            Cell("i", "olt", tr("OLT MA5683T")),
            Cell("i", "connector", tr("Pass-through connectors")),
            Cell("i", "pt", tr("Pigtails")),
            Cell("i", "ups", tr("UPS 1500 or 3000")),
            Cell("s", "deleted_ts", tr("Deleted"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("User"), readonly=True)
        ]

        self.OBJECT_CELLS = [
            Cell("s", "name", tr("Object"), validate=self.validate_object),
            Cell("s", "addr", tr("Address")),
            Cell("i", "blocks", tr("Blocks"), 1, 25),
            Cell("i", "levels", tr("Levels")),
            Cell("i", "apartments", tr("Apartments")),
            Cell("s", "branch", tr("Branch")),
            Cell("s", "comment", tr("Comment")),
            Cell("s", "extra_codes", tr("Mark")),
            Cell("s", "deleted_ts", tr("Deleted"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("User"), readonly=True)
        ]

        self.ITEM_CELLS_PREFIX = [
            Cell("e", "name", tr("Equipment and supplies"), validate=self.validate_item, readonly=not backend.user["advanced_editor"]),
            Cell("i", "total", tr("Total"), calculated=True)
        ]
        self.ITEM_CELLS_SUFFIX = [
            Cell("i", "basement", tr("Basement")),
            Cell("i", "project", tr("Project"), readonly=not backend.user["advanced_editor"]),
            Cell("i", "reserved", tr("Reserved"), calculated=True),
            Cell("s", "comment", tr("Comment")),
            Cell("s", "deleted_ts", tr("Deleted"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("User"), readonly=True)
        ]

        self.CABLE_CELLS_PREFIX = [
            Cell("s", "name", tr("Mark"), validate=self.validate_cable, readonly=not backend.user["advanced_editor"]),
            Cell("f", "total", tr("Total"), calculated=True)
        ]
        self.CABLE_CELLS_SUFFIX = [
            Cell("f", "basement", tr("Basement")),
            Cell("f", "project", tr("Project"), readonly=not backend.user["advanced_editor"]),
            Cell("f", "reserved", tr("Reserved"), calculated=True),
            Cell("s", "comment", tr("Comment")),
            Cell("s", "deleted_ts", tr("Deleted"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("User"), readonly=True)
        ]

        self.olt.currentTextChanged.connect(self.olt_changed)
        self.__olt_props = Table(self.olt_props, self.OLT_CELLS, [], None, backend, "olt", self.on_create_editor)
        self.project.currentTextChanged.connect(self.project_changed)
        self.__objects = Table(self.objects, self.OBJECT_CELLS, [], self.object_calculate, backend, "object", self.on_create_editor)
        self.objects.selectionModel().selectionChanged.connect(self.object_selection_changed)
        self.__items = None
        self.__cables = None

        reload_combobox(olts, self.olt, self.olt_changed)

        self.olt_insert.clicked.connect(self.olt_insert_clicked)
        self.olt_delete.clicked.connect(self.olt_delete_clicked)

        self.olt_rename = QAction(tr("Rename"), self)
        self.olt_rename.triggered.connect(self.olt_rename_clicked)
        self.olt_actions.addAction(self.olt_rename)

        separator = QAction(self)
        separator.setSeparator(True)
        self.olt_actions.addAction(separator)

        self.olt_history = QAction(tr("History"), self)
        self.olt_history.triggered.connect(self.olt_history_clicked)
        self.olt_actions.addAction(self.olt_history)

        separator = QAction(self)
        separator.setSeparator(True)
        self.olt_actions.addAction(separator)

        self.olt_report = QAction(tr("Report"), self)
        self.olt_report.triggered.connect(self.olt_report_clicked)
        self.olt_actions.addAction(self.olt_report)

        self.project_insert.clicked.connect(self.project_insert_clicked)
        self.project_delete.clicked.connect(self.project_delete_clicked)
        self.project_comment.clicked.connect(self.project_comment_clicked)
        self.project_load.clicked.connect(self.project_load_clicked)

        self.project_rename = QAction(tr("Rename"), self)
        self.project_rename.triggered.connect(self.project_rename_clicked)
        self.project_actions.addAction(self.project_rename)

        separator = QAction(self)
        separator.setSeparator(True)
        self.project_actions.addAction(separator)

        self.project_history = QAction(tr("History"), self)
        self.project_history.triggered.connect(self.project_history_clicked)
        self.project_actions.addAction(self.project_history)

        self.object_insert.clicked.connect(self.object_insert_clicked)
        self.item_insert.clicked.connect(self.item_insert_clicked)
        self.cable_insert.clicked.connect(self.cable_insert_clicked)

        self.object_delete.clicked.connect(self.object_delete_clicked)
        self.item_delete.clicked.connect(self.item_delete_clicked)
        self.cable_delete.clicked.connect(self.cable_delete_clicked)

        self.history.clicked.connect(self.history_clicked)
        self.user_history.clicked.connect(self.user_history_clicked)
        self.users.clicked.connect(self.users_clicked)
        self.set_password.clicked.connect(self.set_password_clicked)
        self.item_types.clicked.connect(self.item_types_clicked)

        self.olt_history.setEnabled(backend.user["admin"])
        self.project_history.setEnabled(backend.user["admin"])

        self.show_deleted.setEnabled(backend.user["admin"])
        self.history.setEnabled(backend.user["admin"])
        self.user_history.setEnabled(backend.user["admin"])
        self.users.setEnabled(backend.user["admin"])

        self.item_types.setEnabled(backend.user["admin"])

        self.olt_insert.setEnabled(backend.user["advanced_editor"])
        self.olt_delete.setEnabled(backend.user["advanced_editor"])
        self.olt_rename.setEnabled(backend.user["advanced_editor"])

        self.project_insert.setEnabled(backend.user["advanced_editor"])
        self.project_delete.setEnabled(backend.user["advanced_editor"])
        self.project_comment.setEnabled(backend.user["editor"] or backend.user["advanced_editor"])
        self.project_load.setEnabled(backend.user["advanced_editor"])
        self.project_rename.setEnabled(backend.user["advanced_editor"])

        self.object_insert.setEnabled(backend.user["advanced_editor"])
        self.object_delete.setEnabled(backend.user["advanced_editor"])

        self.item_insert.setEnabled(backend.user["advanced_editor"])
        self.item_delete.setEnabled(backend.user["advanced_editor"])

        self.cable_insert.setEnabled(backend.user["advanced_editor"])
        self.cable_delete.setEnabled(backend.user["advanced_editor"])

        self.active_table = None
        self.active_row = None
        self.active_column = None
        self.active_row_name = None

        self.show_deleted.stateChanged.connect(self.show_deleted_changed)

        if backend.user["admin"]:
            f1 = QAction(self)
            f1.setShortcut(Qt.Key_F1)
            f1.triggered.connect(self.history_clicked)
            self.addAction(f1)
        self.history.setText(self.history.text() + " (F1)")

    def olt_changed(self, _):
        id = self.olt.currentData()

        olt_props = backend.select("olt", {"id": id})
        if olt_props is False:
            return
        self.__olt_props.reload(olt_props, self.show_deleted.isChecked(), not backend.user["advanced_editor"])

        projects = backend.select("project", {"olt_id": id}, "name")
        if projects is False:
            return
        reload_combobox(projects, self.project, self.project_changed, self.show_deleted.isChecked())

        if not self.project.count():
            self.objects.setRowCount(0)
            self.items.setRowCount(0)
            self.cables.setRowCount(0)

    def project_changed(self, _):
        id = self.project.currentData()

        objects = backend.select("object", {"project_id": id}, "name")
        if objects is False:
            return
        self.__objects.reload(objects, self.show_deleted.isChecked(), not backend.user["advanced_editor"])

        self.items.setRowCount(0)
        self.cables.setRowCount(0)

    def object_selection_changed(self, selected, deselected):
        selected = selected.indexes()
        deselected = deselected.indexes()
        if not selected or (selected and deselected and selected[0].row() == deselected[0].row()):
            return
        obj = self.__objects.selected_row()
        self.reload_items(obj)
        self.reload_cables(obj)

    def olt_insert_clicked(self):
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("OLT"), QLineEdit.Normal, tr("New OLT"))
        if not name or not ok:
            return
        if self.olt.findText(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        olt = backend.insert("olt", {"name": name})
        if olt is False:
            return
        self.olt.addItem(olt["name"], olt["id"])
        self.olt.setCurrentIndex(self.olt.count() - 1)

    def project_insert_clicked(self):
        if not self.olt.count():
            return
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("Project"), QLineEdit.Normal, tr("New project"))
        name = name.replace("*", "")
        if not name or not ok:
            return
        if self.project.findText("^{}\**$".format(re.escape(name)), Qt.MatchRegularExpression) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        olt_id = self.olt.currentData()
        project = backend.insert("project", {"name": name, "olt_id": olt_id})
        if project is False:
            return
        self.project.addItem(project["name"], project["id"])
        self.project.setCurrentIndex(self.project.count() - 1)

    def object_insert_clicked(self):
        if not self.project.count():
            return
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("Object"), QLineEdit.Normal, tr("New object"))
        if not name or not ok:
            return
        if self.__objects.find(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        project_id = self.project.currentData()
        obj = backend.insert("object", {"name": name, "project_id": project_id})
        if obj is False:
            return
        self.__objects.insert(obj)

    def item_insert_clicked(self):
        obj = self.__objects.selected_row()
        if obj is None:
            return
        name, ok = QInputDialog().getItem(self, tr("Insert"), tr("Equipment and supplies"), item_type_names, editable=False)
        if not name or not ok:
            return
        if self.__items.find(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        item = backend.insert("item", {"name": name, "object_id": obj.get_id()})
        if item is False:
            return
        self.__items.insert(item)

    def cable_insert_clicked(self):
        obj = self.__objects.selected_row()
        if obj is None:
            return
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("Mark"), QLineEdit.Normal, "")
        if not name or not ok:
            return
        if cable_type(name) is None:
            QMessageBox.warning(self, tr("Warning"), tr("Use a cable type such as FO24 as a prefix."))
            return
        if self.__cables.find(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        cable = backend.insert("cable", {"name": name, "object_id": obj.get_id()})
        if cable is False:
            return
        self.__cables.insert(cable)

    def olt_rename_clicked(self):
        if not self.olt.count():
            return
        id = self.olt.currentData()
        name, ok = QInputDialog().getText(self, tr("Rename"), tr("OLT"), QLineEdit.Normal, self.olt.currentText())
        if not name or not ok:
            return
        index = self.olt.findText(name)
        if index != -1 and index != self.olt.currentIndex():
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        if backend.update("olt", {"id": id, "name": name}) is False:
            return
        self.olt.setItemText(self.olt.currentIndex(), name)

    def project_rename_clicked(self):
        if not self.project.count():
            return
        id = self.project.currentData()
        name, ok = QInputDialog().getText(self, tr("Rename"), tr("Project"), QLineEdit.Normal, self.project.currentText().replace("*", ""))
        name = name.replace("*", "")
        if not name or not ok:
            return
        index = self.project.findText("^{}\**$".format(re.escape(name)), Qt.MatchRegularExpression)
        if index != -1 and index != self.project.currentIndex():
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        if backend.update("project", {"id": id, "name": name}) is False:
            return
        has_comment = self.project.currentText().find("*") != -1
        self.project.setItemText(self.project.currentIndex(), "{}***".format(name) if has_comment else name)

    def olt_delete_clicked(self):
        if not self.olt.count():
            return
        if danger(self, tr("Warning"), tr("Delete OLT?"), QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        id = self.olt.currentData()
        if backend.delete("olt", id) is False:
            return
        self.olt.removeItem(self.olt.currentIndex())
        if not self.olt.count():
            self.project.clear()
            self.objects.setRowCount(0)
            self.items.setRowCount(0)
            self.cables.setRowCount(0)

    def project_delete_clicked(self):
        if not self.project.count():
            return
        if danger(self, tr("Warning"), tr("Delete project?"), QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        id = self.project.currentData()
        if backend.delete("project", id) is False:
            return
        self.project.removeItem(self.project.currentIndex())
        if not self.project.count():
            self.objects.setRowCount(0)
            self.items.setRowCount(0)
            self.cables.setRowCount(0)

    def object_delete_clicked(self):
        obj = self.__objects.selected_row()
        if obj is None:
            return
        if danger(self, tr("Warning"), tr("Delete object?"), QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        if backend.delete("object", obj.get_id()) is False:
            return
        self.objects.removeRow(obj.get_row())
        if not self.objects.rowCount():
            self.items.setRowCount(0)
            self.cables.setRowCount(0)

    def item_delete_clicked(self):
        obj = self.__objects.selected_row()
        if obj is None:
            return
        item = self.__items.selected_row()
        if item is None:
            return
        if danger(self, tr("Warning"), tr("Delete row?"), QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        if backend.delete("item", item.get_id()) is False:
            return
        self.items.removeRow(item.get_row())

    def cable_delete_clicked(self):
        obj = self.__objects.selected_row()
        if obj is None:
            return
        cable = self.__cables.selected_row()
        if cable is None:
            return
        if danger(self, tr("Warning"), tr("Delete row?"), QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        if backend.delete("cable", cable.get_id()) is False:
            return
        self.cables.removeRow(cable.get_row())

    def olt_report_clicked(self):
        fn, _ = QFileDialog.getSaveFileName(self, filter="XLSX (*.xlsx)")
        if not fn:
            return
        result = backend.get_olt_report(self.olt.currentData())
        if result is False:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, tr("OLT"))
        ws.cell(1, 2, self.olt.currentText())
        font = copy(ws.cell(1, 1).font)
        font.bold = True
        fill = PatternFill("solid", start_color="dddddd")
        for row, cells in enumerate(result):
            ws.cell(row + 2, 1, cells[0])
            column = 1
            while column < len(cells):
                cell = cells[column]
                ws.cell(row + 2, 2 * column, cell[0]).font = font
                ws.cell(row + 2, 2 * column + 1, cell[1]).fill = fill
                column += 1
        column = 1
        while column < len(result[0]):
            ws.cell(2, 2 * column + 1, tr("Project"))
            column += 1
        ws.cell(2, 1, "{}\\{}".format(tr("Equipment and supplies"), tr("Project")))
        wb.save(fn)
        wb.close()
        QMessageBox.information(self, tr("Info"), tr("Success."))

    def olt_history_clicked(self):
        global history, history_window
        filter_by = {
            "table": "olt",
            "row": self.olt.currentData(),
            "column": "name"
        }
        history = backend.select("history", filter_by, "id")
        if history is False:
            return
        history_window = HistoryWindow(self, **filter_by, row_name=self.olt.currentText())
        history_window.show()

    def project_history_clicked(self):
        global history, history_window
        item_to_column = {
            to_widget_column("name"): "name",
            to_widget_column("comment"): "comment"
        }
        item, ok = QInputDialog().getItem(self, tr("History"), tr("Column"), item_to_column.keys(), editable=False)
        if not ok:
            return
        filter_by = {
            "table": "project",
            "row": self.project.currentData(),
            "column": item_to_column[item]
        }
        history = backend.select("history", filter_by, "id")
        if history is False:
            return
        history_window = HistoryWindow(self, **filter_by, row_name=self.project.currentText().replace("*", ""))
        history_window.show()

    def history_clicked(self):
        global history, history_window
        if self.active_table is None:
            return
        filter_by = {
            "table": self.active_table,
            "row": self.active_row,
            "column": self.active_column
        }
        history = backend.select("history", filter_by, "id")
        if history is False:
            return
        history_window = HistoryWindow(self, **filter_by, row_name=self.olt.currentText() if self.active_table == "olt" else self.active_row_name)
        history_window.show()

    def user_history_clicked(self):
        global users, user_history_window
        users = backend.select("user", None, "id")
        if users is False:
            return
        user_history_window = UserHistoryWindow(self)
        user_history_window.show()

    def users_clicked(self):
        global users, users_window
        users = backend.select("user", None, "id")
        if users is False:
            return
        users_window = UsersWindow(self)
        users_window.show()

    def set_password_clicked(self):
        password, ok = QInputDialog().getText(self, tr("Change"), tr("Password"), QLineEdit.Password, "")
        if not ok:
            return
        result = backend.set_password(backend.name, password)
        (QMessageBox.information if result else QMessageBox.critical)(
            self,
            tr("Info") if result else tr("Error"),
            tr("Password changed.") if result else tr("Password change error.")
        )

    def item_types_clicked(self):
        global item_types_window
        item_types_window = ItemTypesWindow(self)
        item_types_window.show()

    def show_deleted_changed(self, state):
        global olts

        olts = backend.select("olt", None, "name")
        if olts is False:
            return
        self.project.clear()
        reload_combobox(olts, self.olt, self.olt_changed, state)

        self.olt_insert.setEnabled(not state and backend.user["advanced_editor"])
        self.olt_delete.setEnabled(not state and backend.user["advanced_editor"])
        self.olt_rename.setEnabled(not state and backend.user["advanced_editor"])

        self.project_insert.setEnabled(not state and backend.user["advanced_editor"])
        self.project_delete.setEnabled(not state and backend.user["advanced_editor"])
        self.project_comment.setEnabled(not state and (backend.user["editor"] or backend.user["advanced_editor"]))
        self.project_load.setEnabled(not state and backend.user["advanced_editor"])

        self.object_insert.setEnabled(not state and backend.user["advanced_editor"])
        self.object_delete.setEnabled(not state and backend.user["advanced_editor"])

        self.item_insert.setEnabled(not state and backend.user["advanced_editor"])
        self.item_delete.setEnabled(not state and backend.user["advanced_editor"])

        self.cable_insert.setEnabled(not state and backend.user["advanced_editor"])
        self.cable_delete.setEnabled(not state and backend.user["advanced_editor"])

    def project_comment_clicked(self):
        if not self.project.count():
            return

        id = self.project.currentData()

        project_props = backend.select("project", {"id": id})
        if project_props is False or not project_props:
            return
        project_props = project_props[0]

        comment, ok = QInputDialog().getMultiLineText(self, tr("Change"), tr("Comment"), project_props["comment"])
        if not ok:
            return

        if backend.update("project", {"id": id, "comment": comment}) is False:
            return

        self.project.setItemText(self.project.currentIndex(), "{}***".format(project_props["name"]) if comment else project_props["name"])

    def project_load_clicked(self):
        if not self.olt.count():
            return
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("Project"), QLineEdit.Normal, tr("New project"))
        name = name.replace("*", "")
        if not name or not ok:
            return
        if self.project.findText("^{}\**$".format(re.escape(name)), Qt.MatchRegularExpression) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        fn, _ = QFileDialog.getOpenFileName(self, filter="XLSX (*.xlsx *.xlsm)")
        if not fn:
            return
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            objects = load(fn)
            QApplication.restoreOverrideCursor()
        except:
            logging.warning("load error: {}".format(traceback.format_exc()))
            QMessageBox.warning(self, tr("Warning"), tr("Error"))
            return
        olt_id = self.olt.currentData()
        project = backend.load_project({"name": name, "olt_id": olt_id, "objects": objects})
        if project is False:
            return
        self.project.addItem(project["name"], project["id"])
        self.project.setCurrentIndex(self.project.count() - 1)
        QMessageBox.information(self, tr("Info"), tr("Success."))


class ItemTypesWindow(QMainWindow, Ui_ItemTypesWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setupUi(self)

        self.ITEM_TYPE_CELLS = [
            Cell("s", "id", tr("Name"), readonly=True)
        ]

        self.__item_types = Table(self.item_types, self.ITEM_TYPE_CELLS, item_types, None, backend, "item_type")

        self.item_type_insert.clicked.connect(self.item_type_insert_clicked)

        esc = QAction(self)
        esc.setShortcut(Qt.Key_Escape)
        esc.triggered.connect(self.close)
        self.addAction(esc)

    def item_type_insert_clicked(self):
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("Name"), QLineEdit.Normal, "")
        if not name or not ok:
            return
        if self.__item_types.find(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        item_type = backend.insert("item_type", {"id": name})
        if item_type is False:
            return
        item_types.append(item_type)
        item_type_names.append(name)
        self.__item_types.insert(item_type)


class UsersWindow(QMainWindow, Ui_UsersWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setupUi(self)

        self.USER_CELLS = [
            Cell("s", "id", tr("Name"), readonly=True),
            Cell("i", "admin", tr("Admin"), min=0, max=1),
            Cell("i", "editor", tr("Editor"), min=0, max=1),
            Cell("i", "advanced_editor", tr("Advanced editor"), min=0, max=1)
        ]

        self.__users = Table(self.users, self.USER_CELLS, users, None, backend, "user")

        self.user_insert.clicked.connect(self.user_insert_clicked)
        self.user_delete.clicked.connect(self.user_delete_clicked)
        self.user_set_password.clicked.connect(self.user_set_password_clicked)

        esc = QAction(self)
        esc.setShortcut(Qt.Key_Escape)
        esc.triggered.connect(self.close)
        self.addAction(esc)

    def user_insert_clicked(self):
        name, ok = QInputDialog().getText(self, tr("Insert"), tr("User"), QLineEdit.Normal, "")
        if not name or not ok:
            return
        if self.__users.find(name) != -1:
            QMessageBox.warning(self, tr("Warning"), tr("Already exists."))
            return
        user = backend.insert("user", {"id": name, "password": "", "admin": 0, "editor": 0, "advanced_editor": 0})
        if user is False:
            return
        users.append(user)
        self.__users.insert(user)

    def user_delete_clicked(self):
        global users
        user = self.__users.selected_row()
        if user is None:
            return
        id = user.get_id()
        if backend.delete("user", id) is False:
            return
        users = [u for u in users if u["id"] != id]
        self.users.removeRow(user.get_row())

    def user_set_password_clicked(self):
        user = self.__users.selected_row()
        if user is None:
            return
        password, ok = QInputDialog().getText(self, tr("Change"), tr("Password"), QLineEdit.Password, "")
        if not ok:
            return
        result = backend.set_password(user.get_id(), password)
        (QMessageBox.information if result else QMessageBox.critical)(
            self,
            tr("Info") if result else tr("Error"),
            tr("Password changed.") if result else tr("Password change error.")
        )


all_tables = {}


def to_widget_table(table):
    if not all_tables:
        all_tables.update({
            "olt": tr("OLT"),
            "project": tr("Projects"),
            "object": tr("Objects"),
            "item": tr("Equipment and supplies"),
            "cable": tr("Cables")
        })
    return all_tables.get(table, table)


all_columns = {}


def to_widget_column(column):
    if not all_columns:
        all_columns.update({
            "gpon": tr("GPON-card x16"),
            "odf": tr("ODF x24"),
            "olt": tr("OLT MA5683T"),
            "connector": tr("Pass-through connectors"),
            "pt": tr("Pigtails"),
            "ups": tr("UPS 1500 or 3000"),
            "deleted_ts": tr("Deleted"),
            "user_id": tr("User"),

            "name": tr("Name"),
            "addr": tr("Address"),
            "blocks": tr("Blocks"),
            "levels": tr("Levels"),
            "apartments": tr("Apartments"),
            "branch": tr("Branch"),
            "comment": tr("Comment"),
            "extra_codes": tr("Mark"),

            "total": tr("Total"),
            "basement": tr("Basement"),
            "project": tr("Project"),
            "reserved": tr("Reserved")
        })
        for i in range(25):
            all_columns["block_{}".format(i + 1)] = str(i + 1)
    return all_columns.get(column, column)


all_actions = {}


def to_widget_action(action):
    if not all_actions:
        all_actions.update({
            1: tr("Edited"),
            2: tr("Deleted")
        })
    return all_actions.get(action, action)


class HistoryWindow(QMainWindow, Ui_HistoryWindow):
    def __init__(self, parent, table, row, column, row_name):
        super().__init__(parent)
        self.setupUi(self)

        self.setWindowTitle(
            "{history_label}: {table_label} - {table}, {row_label} - {row}, {column_label} - {column}".format(
                history_label=tr("History"),
                table_label=tr("Table"), table=to_widget_table(table),
                row_label=tr("Row"), row=row_name,
                column_label=tr("Column"), column=to_widget_column(column)
            )
        )

        self.HISTORY_CELLS = [
            Cell("s", "ts", tr("Editing time"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("Editor"), readonly=True),
            Cell("s", "value", tr("Value"), readonly=True)
        ]

        self.__history = Table(self.history, self.HISTORY_CELLS, history)

        self.setFocus()

        esc = QAction(self)
        esc.setShortcut(Qt.Key_Escape)
        esc.triggered.connect(self.close)
        self.addAction(esc)


class SelectUsersWindow(QMainWindow, Ui_SelectUsersWindow):
    def __init__(self, parent, selected_users, receive_selected_users):
        super().__init__(parent)
        self.setupUi(self)

        for user in users:
            item = QListWidgetItem(user["id"], self.users)
            item.setCheckState(Qt.Checked if user["id"] in selected_users else Qt.Unchecked)
            self.users.addItem(item)

        self.receive_selected_users = receive_selected_users

        self.ok.clicked.connect(self.ok_clicked)

        esc = QAction(self)
        esc.setShortcut(Qt.Key_Escape)
        esc.triggered.connect(self.close)
        self.addAction(esc)

    def ok_clicked(self):
        user_ids = []
        for row in range(self.users.count()):
            item = self.users.item(row)
            if item.checkState() == Qt.Checked:
                user_ids.append(item.text())
        self.receive_selected_users(user_ids)
        self.close()


class UserHistoryWindow(QMainWindow, Ui_UserHistoryWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setupUi(self)

        self.HISTORY_CELLS = [
            Cell("s", "ts", tr("Editing time"), readonly=True, to_widget=to_widget_ts),
            Cell("s", "user_id", tr("Editor"), readonly=True),
            Cell("s", "olt", tr("OLT"), readonly=True),
            Cell("s", "project", tr("Project"), readonly=True),
            Cell("s", "object", tr("Object"), readonly=True),
            Cell("s", "table", tr("Table"), readonly=True, to_widget=to_widget_table),
            Cell("s", "row_name", tr("Row"), readonly=True),
            Cell("s", "column", tr("Column"), readonly=True, to_widget=to_widget_column),
            Cell("s", "value", tr("Value"), readonly=True),
            Cell("s", "action", tr("Action"), readonly=True, to_widget=to_widget_action)
        ]

        self.calendar.setLocale(QLocale(settings["language"]))
        self.calendar.selectionChanged.connect(self.calendar_selection_changed)
        self.calendar.clicked.connect(self.calendar_clicked)
        self.active_start_of_day = None

        self.select_users.clicked.connect(self.select_users_clicked)

        self.page.currentTextChanged.connect(self.get_user_history)
        self.prev_page.clicked.connect(self.prev_page_clicked)
        self.next_page.clicked.connect(self.next_page_clicked)

        self.__history = Table(self.history, self.HISTORY_CELLS, [])

        self.setFocus()

        esc = QAction(self)
        esc.setShortcut(Qt.Key_Escape)
        esc.triggered.connect(self.close)
        self.addAction(esc)

    def get_selected_users(self):
        user_ids = []
        for row in range(self.users.count()):
            user_ids.append(self.users.item(row).text())
        return user_ids

    def get_user_history(self):
        if self.active_start_of_day is None:
            return
        page = self.page.currentIndex()
        result = backend.get_user_history(self.get_selected_users(), self.active_start_of_day, page)
        if result is False:
            return
        if self.page.count() != result["chunk_count"]:
            self.page.currentTextChanged.disconnect()
            self.page.clear()
            for i in range(result["chunk_count"]):
                self.page.addItem(str(i + 1))
            self.page.setCurrentIndex(min(page, result["chunk_count"] - 1))
            self.page.currentTextChanged.connect(self.get_user_history)
        self.__history.reload(result["chunk"])

    def receive_selected_users(self, user_ids):
        self.users.clear()
        for user_id in user_ids:
            self.users.addItem(user_id)
        self.get_user_history()

    def select_users_clicked(self):
        global select_users_window
        select_users_window = SelectUsersWindow(self, self.get_selected_users(), self.receive_selected_users)
        select_users_window.show()

    def calendar_selection_changed(self):
        self.active_start_of_day = self.calendar.selectedDate().startOfDay().toSecsSinceEpoch()
        self.page.currentTextChanged.disconnect()
        self.page.clear()
        self.page.addItem("1")
        self.page.currentTextChanged.connect(self.get_user_history)
        self.get_user_history()

    def calendar_clicked(self, date):
        if self.active_start_of_day is None:
            self.active_start_of_day = self.calendar.selectedDate().startOfDay().toSecsSinceEpoch()
            self.get_user_history()

    def prev_page_clicked(self):
        self.page.setCurrentIndex(max(self.page.currentIndex() - 1, 0))

    def next_page_clicked(self):
        self.page.setCurrentIndex(min(self.page.currentIndex() + 1, self.page.count() - 1))


qt_translator = None
qt_base_translator = None
app_translator = None


def set_language(language):
    global qt_translator, qt_base_translator, app_translator

    with open(os.path.join(app_wd, "settings.json"), "wb") as f:
        settings.update({
            "language": language
        })
        f.write(json.dumps(settings).encode("UTF-8"))

    locale = QLocale(language)
    translations_dir = QLibraryInfo.path(QLibraryInfo.TranslationsPath)

    if qt_translator is not None:
        app.removeTranslator(qt_translator)
    qt_translator = QTranslator()
    if qt_translator.load(locale, "qt", "_", translations_dir):
        app.installTranslator(qt_translator)
    else:
        qt_translator = None

    if qt_base_translator is not None:
        app.removeTranslator(qt_base_translator)
    qt_base_translator = QTranslator()
    if qt_base_translator.load(locale, "qtbase", "_", translations_dir):
        app.installTranslator(qt_base_translator)
    else:
        qt_base_translator = None

    if app_translator is not None:
        app.removeTranslator(app_translator)
    app_translator = QTranslator()
    if app_translator.load(locale, "app", "_", ":/i18n"):
        app.installTranslator(app_translator)
    else:
        app_translator = None


if __name__ == "__main__":
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("fo.db")
    app = QApplication(sys.argv)

    set_language(settings["language"])

    enter_window = EnterWindow()
    enter_window.show()

    os._exit(app.exec())

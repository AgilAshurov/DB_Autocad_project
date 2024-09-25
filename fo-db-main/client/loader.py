import openpyxl
import re
from startup import app_wd
import os


AT_VISIBILITY = 3
AT_BLOCKNAME = 4
AT_CODE = 8
AT_NUMBER = 10
AT_ADDR = 11
AT_LEVELS = 12
AT_BLOCKS = 13
AT_APARTMENTS = 14
AT_FO = 16
AT_FO_L = 17
AT_FO_MTK = 18
AT_FO_MTK_L = 19
AT_FO2_DC_L1 = 24
AT_FO2_DC_L2 = 25
AT_EXTRA_CODE = 28

SPLITTER_TYPE_RE = re.compile("^x([0-9]+)$")
PILLAR_TYPE_RE = re.compile("^STolb_([0-9]+)$")
CABINET_TYPE_RE = re.compile("^.*_x([0-9]+)$")
WIDE_CODE_RE = re.compile("^[A-Z]{2}")


def extract_type(e, s):
    m = e.match(s)
    return None if m is None else m.group(1)


def convert(value, t, default):
    return default if value == " " or value == "" or value is None else t(value)


def read(ws, r, c, t=str, default=""):
    return convert(ws.cell(r, c).value, t, default)


def load(fn):
    objects = []
    current_code = None
    wide_code = False
    obj = None
    wb = openpyxl.load_workbook(fn, data_only=True)
    ws = wb["AT"]
    for row in range(7, ws.max_row):
        visibility = read(ws, row, AT_VISIBILITY)
        blockname = read(ws, row, AT_BLOCKNAME)
        code = read(ws, row, AT_CODE)
        fo = read(ws, row, AT_FO)
        fo_mtk = read(ws, row, AT_FO_MTK)

        if not code:
            continue

        if current_code != code:
            wide_code = WIDE_CODE_RE.match(code) is not None
            obj = {
                "comment": code,
                "branch": code[0] if code else "",
                "name": code,
                "addr": "",
                "levels": 0,
                "blocks": 1,
                "apartments": 0,
                "items": {},
                "cables": {},
                "cabinet_count": 0,
                "extra_codes": []
            }
            objects.append(obj)
            current_code = code

        if wide_code:
            extra_code = read(ws, row, AT_EXTRA_CODE)
            if extra_code:
                obj["extra_codes"].append(extra_code)

        if blockname == ".Dom":
            obj.update({
                "name": read(ws, row, AT_NUMBER),
                "addr": read(ws, row, AT_ADDR),
                "levels": read(ws, row, AT_LEVELS, int, 0),
                "blocks": read(ws, row, AT_BLOCKS, int, 0),
                "apartments": read(ws, row, AT_APARTMENTS, int, 0)
            })

        if blockname.startswith("ОРШ_") or blockname.startswith("ВРШ_"):
            obj["cabinet_count"] += 1

        for item_id_template, e, s in [
            ("Cartridge splitter  x{}", SPLITTER_TYPE_RE, visibility),
            ("Stolb  {}m", PILLAR_TYPE_RE, blockname),
            ("Şkaf  x{}", CABINET_TYPE_RE, blockname)
        ]:
            t = extract_type(e, s)
            if t:
                item_id = item_id_template.format(t)
                obj["items"][item_id] = obj["items"].get(item_id, 0) + 1

        if blockname == ".FO2_DC":
            obj["cables"]["FO2_DC"] = obj["cables"].get("FO2_DC", 0) + read(ws, row, AT_FO2_DC_L1, int, 0) + read(ws, row, AT_FO2_DC_L2, int, 0)

        if blockname == "Муфта":
            item_id = "Mufta  1x48"
            obj["items"][item_id] = obj["items"].get(item_id, 0) + 1

        if fo:
            obj["cables"][fo] = read(ws, row, AT_FO_L, int, 0)
        if fo_mtk:
            obj["cables"][fo_mtk] = read(ws, row, AT_FO_MTK_L, int, 0)
    wb.close()
    for obj in objects:
        tubes = (obj["levels"] - 2) * obj["blocks"] if obj["levels"] > 2 else 0
        cabinet_count = obj.pop("cabinet_count")
        if tubes:
            obj["items"]["Mərtəbə arası Borular 2,5m (əd)"] = tubes
        if cabinet_count:
            obj["items"]["Şlanq  Plasmas  d20"] = cabinet_count * 2
            obj["items"]["Trank  60x60mm"] = cabinet_count
        obj["items"]["Trank  40x25mm"] = obj["blocks"] * 2
        obj["extra_codes"].sort()
        obj["extra_codes"] = " ".join(obj["extra_codes"])
        for cable_id, cable_length in obj["cables"].items():
            if not cable_length:
                raise RuntimeError("bad cable length for {}".format(cable_id))
    return objects


if __name__ == "__main__":
    objects = load(os.path.join(app_wd, "test.xlsx"))
    for obj in objects:
        print("-" * 70)
        for k, v in obj.items():
            if isinstance(v, dict):
                for sk, sv in v.items():
                    print(sk, ":", sv)
            else:
                print(k, ":", v)
